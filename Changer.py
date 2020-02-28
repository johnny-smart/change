import openpyxl
import csv
import config
from profilehooks import timecall, profile
# from geolocation import geoadressation
from os import remove
from os import path
from err_decorator import error_module
from functools import partial
import json
from agregate import check_changer

err_rec = []
# geodata = []


double = []


# функция для создания и заполнения массива записей домов
# (ввод: имя внутри функции, вывод: маcсив записей)
# @timecall
# def houses_init():
#     houses = []
#     houses_wb = openpyxl.load_workbook(config.HOUSES)
#     houses_sn = houses_wb.sheetnames[0]
#     w_sheet = houses_wb[houses_sn]
#     houses_list = w_sheet.rows
#     next(houses_list)
#     for row in houses_list:
#         cols = []
#         for init in row:
#             cols.append(init.value)
#         houses.append(cols)
#     return(houses)


# def houses_filter(town, houses):
#     houses_filtred = []
#     for row in houses:
#         if (row[1] in town):
#             houses_filtred.append(row)
#     return houses_filtred


@timecall
def hardware_init(fname, sheet):

    hardware = []

    hardware_wb = openpyxl.load_workbook(fname)
    hardware_list = hardware_wb[sheet].rows

    next(hardware_list)

    for row in hardware_list:
        cols = [None]*20
        for init in row:
            cols[init.column-1] = str(init.value)

            if (((init.column == 10) or (init.column == 11)) and
               (init.value is not None)):
                cols[init.column-1] = str(init.value).split(' ')[0]
            if (((init.column == 1) or (init.column == 2)) and (init.font.strike is True)):
                cols[19] = 1
            if (init.column == 4 or init.column == 5):
                if(init.fill.fgColor.rgb == 'FF00B0F0'):
                    cols[18] = 1
        if not (cols[0] is None):
            hardware.append(cols)
    return(hardware)


@timecall
def result_init(arg):
    town, sheet = arg[0], arg[1]

    _err = []

    _result = hardware_init(config.HARDWARE, sheet)

    print("hardware", len(_result))
    _result, _err = error_module(_result,  town)
    print("result:", len(_result))
    return {'result':_result, 'err':_err,}


@timecall
def out_file(result, namefile):
    result = [x for i,x in enumerate(result) if not x is None]


    # try:
    if path.isfile(config.DIR + namefile + '.csv'):
        remove(config.DIR + namefile + '.csv')
    with open(

                config.DIR + namefile + '.csv',
                'a+',
                newline=''

            ) as newfile:

        scvwr = csv.writer(newfile, delimiter=';', quoting=csv.QUOTE_MINIMAL)

        scvwr.writerows(result)

def format_founded(excel_lost, map_found, address):
    if address[0] == 'P_STREET':
        return address

    for i, item in enumerate(excel_lost):
        if item[0] == address[0] and item[1] == address[1]:
            return False

    for item in map_found:
        founded_address = map_found[item]['hardware']
        if address[0] == founded_address[0] and address[1] == founded_address[1]:
            address_tmp = address[4].split(';')
            if (address[3]==item or item in address_tmp):
                address = founded_address
                if address[7].find('\n') != -1:
                    address[7] = "|".join(address[7].split('\n'))

                return address
    else:
        print('\n Не найдено соответствие :', address)
        return False



def regions_worker(reg_list, flag_mod=''):
    _err = []

    _result = [['P_STREET', 'P_HOUSE', 'P_MODEL', 'P_IP_OLD', 'P_IP',
            'P_VECTOR', 'P_UPLINK', 'P_MAC', 'P_VLAN', 'P_DATE_SETUP',
            'P_DATE_INSTALL', 'P_FLATS', 'P_DOOR', 'P_FLOOR',
            'P_DESCRIPTION', 'P_RESERVED1', 'P_RESERVED2', 'P_RESERVED3',
            'P_TRANSIT', 'P_REMOVED',]]

    output_init = list(map(result_init,reg_list))

    for item in output_init:
        _result+=(item['result'])
        _err+=(item['err'])

    for i, d in enumerate(_result):
        if d is None:
            print (i)

    map_not_found, found = check_changer()
    _result = [x for i,x in enumerate(_result) if x is not None]
    format_founded_tmp = partial(format_founded, map_not_found, found)

    _result = list(map(format_founded_tmp, _result))

    _removed = [x for i,x in enumerate(_err) if x[19] == 1]

    _result = [x for i,x in enumerate(_result) if x ]

    _err = [x for i,x in enumerate(_err) if x[19] is None]

    _err = list(map(format_founded_tmp, _err))

    _err = [x for i,x in enumerate(_err) if x ]

    _err.insert(0,['P_STREET', 'P_HOUSE', 'P_MODEL', 'P_IP_OLD', 'P_IP',
            'P_VECTOR', 'P_UPLINK', 'P_MAC', 'P_VLAN', 'P_DATE_SETUP',
            'P_DATE_INSTALL', 'P_FLATS', 'P_DOOR', 'P_FLOOR',
            'P_DESCRIPTION', 'P_RESERVED1', 'P_RESERVED2', 'P_RESERVED3',
            'P_TRANSIT', 'P_REMOVED',])

    out_file(_removed, 'Removed' + flag_mod)
    out_file(_result, 'Result' + flag_mod)
    out_file(_err, 'Err' + flag_mod)
    print()


@profile
def main():

    regions_worker(
        [
            [
                'Д. КАБАНОВО',
                'Комутаторы КБ',
            ],
            [
                'Г. ЛИКИНО-ДУЛЕВО',
                'Комутаторы ЛД',
            ],
            [
                'Г. ОРЕХОВО-ЗУЕВО',
                'Комутаторы ОЗ',
            ],
            [
                'Г. КУРОВСКОЕ',
                'Комутаторы КУ',
            ],
            [
                ['Д. ДЕМИХОВО', 'Д. НАЖИЦЫ', 'Д. КРАСНАЯ ДУБРАВА'],
                'Комутаторы ДМ',
            ],
        ],

        )


    # regions_worker(
    #     [[
    #             'Г. ОРЕХОВО-ЗУЕВО',
    #             'Комутаторы ОЗ',
    #     ]],

    #     '_OZ'
    #     )
    print('done')


if __name__ == "__main__":

    main()
